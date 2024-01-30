import requests
import chardet
from bs4 import BeautifulSoup
import openpyxl


def check_https(url, timeout=10):
    try:
        response = requests.head(url, timeout=timeout)
        if response.status_code == requests.codes.ok:
            if response.url.startswith("https"):
                return True
    except requests.exceptions.RequestException:
        pass
    return False

def get_body(url, timeout):
    response = requests.get(url=url, timeout=timeout)
    encoding = chardet.detect(response.content)['encoding']
    response.encoding = encoding

    # Bây giờ nội dung sẽ được mã hóa bằng UTF-8
    content = response.text
    soup = BeautifulSoup(content, "html.parser")

    # Extract body content
    body_content = soup.find('body').get_text() if soup.find('body') else None

    print("---------------------------------------------------------------------------")
    print("URL:", url)
    print("Body Content:", body_content.strip())

    data_to_append = [url, body_content.strip()]
    file_path = '.\\Data\\out_with_body.xlsx'
    append_to_excel(file_path=file_path, data=data_to_append)

def append_to_excel(file_path, data):
    try:
        # Mở tệp Excel đã tồn tại
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # Nếu tệp không tồn tại, tạo mới một workbook
        workbook = openpyxl.Workbook()

    # Chọn sheet mặc định (Sheet 1)
    sheet = workbook.active

    # Xác định dòng mới để thêm
    next_row = sheet.max_row + 1

    # Ghi dữ liệu vào dòng mới
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)

    # Lưu lại tệp Excel
    workbook.save(file_path)


workbook = openpyxl.load_workbook('.\\Data\\NOCVN-Mô tả-TMVP.11.2023.xlsx')
# Select the active sheet
sheet = workbook.active

# Get the values from the second column (column B)
column_index = 1
column_values = []

# Iterate over each row in the sheet
for row in sheet.iter_rows(values_only=True):
    cell_value = row[column_index - 1]  # Adjust the index to 0-based index
    column_values.append(cell_value)

# Print the values from the second column
for value in column_values:
    try:
        url = "http://" + value
        is_https = check_https(url=url, timeout=5)
        if is_https:
            get_body("https://" + value, timeout=5)
        else:
            get_body("http://" + value, timeout=5)
    except Exception:
        continue

# Close the workbook
workbook.close()

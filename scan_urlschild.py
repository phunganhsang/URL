import requests
import chardet
from bs4 import BeautifulSoup
import openpyxl
import re
from openpyxl.styles import Alignment
from urllib.parse import urlparse

def check_https(url,timeout=10):
    try:
        response = requests.head(url,timeout=timeout)
        if response.status_code == requests.codes.ok:
            if response.url.startswith("https"):
                return True
    except requests.exceptions.RequestException:
        pass
    return False

def get_meta(url,timeout):
    response = requests.get(url=url,timeout=timeout)
    encoding = chardet.detect(response.content)['encoding']
    response.encoding = encoding

    # Bây giờ nội dung sẽ được mã hóa bằng UTF-8
    content = response.text

    # Example usage
    text_with_urls = content
    urls = extract_urls(text_with_urls)
    # print()
    # print("Extracted URLs:")
    # for url in urls:
    #     print(url)

    soup = BeautifulSoup(content, "html.parser")

    keyword_metadata = soup.find("meta", attrs={"name": "keywords"})
    description_metadata = soup.find("meta", attrs={"name": "description"})
    author_metadata = soup.find("meta", attrs={"name": "author"})

    keyword = keyword_metadata["content"] if keyword_metadata else None
    description = description_metadata["content"] if description_metadata else None
    author = author_metadata["content"] if author_metadata else None
    print("---------------------------------------------------------------------------")
    print("URL:", url)
    print("Keyword:", keyword)
    print("Description:", description)
    print("Author:", author)


    data_to_append = [url, keyword, description, author, "\n".join(urls)]
    file_path = '.\\Data\\out_with_urlchild.xlsx'
    append_to_excel(file_path=file_path,data=data_to_append)

def append_to_excel(file_path, data):
    try:
        # Mở tệp Excel đã tồn tại
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # Nếu tệp không tồn tại, tạo mới một workbook
        workbook = openpyxl.Workbook()

    # Chọn sheet mặc định (Sheet 1)
    sheet = workbook.active

    # Xác định dòng mới để thêm
    next_row = sheet.max_row + 1

    # Ghi dữ liệu vào dòng mới
    for col_num, value in enumerate(data, start=1):
        cell = sheet.cell(row=next_row, column=col_num, value=value)
        cell.alignment = Alignment(wrap_text=True)
    # Lưu lại tệp Excel
    workbook.save(file_path)

def extract_domain(url):
    parsed_url = urlparse(url)
    return parsed_url.netloc

def extract_urls(text):
    # Define a regex pattern for matching URLs
    url_pattern = re.compile(r'https?://\S+|www\.\S+')

    # Use findall to extract all URLs from the text
    raw_urls = re.findall(url_pattern, text)

    # Use a set to store unique domains
    unique_domains = set()

    # Extract and store unique domains
    for url in raw_urls:
        domain = extract_domain(url)
        unique_domains.add(domain)

    return list(unique_domains)

workbook = openpyxl.load_workbook('.\\Data\\NOCVN-Mô tả-TMVP.11.2023.xlsx')
# Select the active sheet
sheet = workbook.active

# Get the values from the second column (column B)
column_index = 1
column_values = []

# Iterate over each row in the sheet
for row in sheet.iter_rows(values_only=True):
    cell_value = row[column_index - 1]  # Adjust the index to 0-based index
    column_values.append(cell_value)

# Print the values from the second column
for value in column_values:
    try:
        # print(value)
        url = "http://"+value
        is_https = check_https(url=url, timeout=5)
        if is_https:
            # print(f"Trang web {url} sử dụng HTTPS")
            get_meta("https://"+value,timeout=5)
        else:
            # print(f"Trang web {url} không sử dụng HTTPS")
            get_meta("http://"+value,timeout=5)
        # get_meta(value)
    except Exception:
        continue
# Close the workbook
workbook.close()
